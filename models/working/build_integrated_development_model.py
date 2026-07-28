#!/usr/bin/env python3
"""
South Andrews Healthcare and Mobility Hub — derived integrated development model.

STATUS: NONCANONICAL WORKING MODEL (models/working/). This is NOT the gated "model vNext"
described in docs/planning/Document Update Order.md Step 5. It does not resolve OQ-14
(program basis), OQ-15 (office rent), OQ-17 (tax credits) or OQ-18 (exit cap); it declares a
stated basis for each and reports the alternative alongside.

PLANNING-LEVEL MODEL OUTPUT. Not an appraisal, contractor estimate, engineered load letter,
tax opinion, lease forecast, or financing commitment.

DERIVATION AND PROVENANCE
  Geometric basis  Basis B (S3 `Assumptions!D9:D12`, from S8 pp.4-5 massing) — 28,000 GSF
                   ground/parking plate, 24,000 GSF office plate, 360 GSF/stall, 85% efficiency.
                   Stated modeling basis, NOT an adopted program. Basis A (S2, 35,000 SF uniform
                   plate) would change every area, stall and dollar figure. Never mix the two.
  Unit costs       S3 `Assumptions!D29:D43` (2026 WGI-benchmarked) as the stated base.
                   S2 alternatives carried in `s2_alternative_unit_costs` for comparison only.
  Soft/financing   S3 `Assumptions!D45:D53`.
  Revenue          S3 `Assumptions!D55:D65`.
  Electrical       Method preserved from the corrected electrical load workbook
                   (South_Andrews_Electrical_Load_Model_CORRECTED.xlsx, SHA256 f7cfe13a...94b400,
                   NOT held in sources/ — see MA-11). Its 18 W/SF clinical density is used here and
                   conflicts with S3's 5 W/GSF; both are reported.
  Exclusions       AV staging revenue is excluded from every base case per AGENTS.md and CAN-012.
                   Solar/storage tax credits underwritten at 0% per S3 `Energy & Mobility!B21`
                   pending tax counsel (OQ-17 / MB-13).

VALIDATION
  Running this model on the SA-B geometry with parking construct P2 and AV revenue re-added
  reproduces S3 `Financing & Returns!D17` ($2,277,951) to within $107 (rounding). The
  governance-clean ex-AV figure reproduces the repository's independent OQ-22 recomputation
  ($2,151,591) to within $107.

Usage:  python3 models/working/build_integrated_development_model.py
Writes: integrated-development-model.json and (if openpyxl is present)
        South_Andrews_Integrated_Development_Model.xlsx, both in this directory.
No original source file is read for write or modified by this script.
"""
import json, math, os

OUT = os.path.dirname(os.path.abspath(__file__))

# ---------------------------------------------------------------- inputs
SITE_SF = 38207
LAND_BASE = 8_000_000          # D-P1 working opening-offer input. Not value/ceiling/authority.
CLOSING = 0.02

PLATE_GROUND = 28_000
PLATE_PARK = 28_000
PLATE_OFFICE = 24_000
GSF_PER_STALL = 360
OFFICE_EFF = 0.85
STALLS_PER_LEVEL = int(PLATE_PARK // GSF_PER_STALL)   # 77

# S3 unit costs (stated base)
U = dict(
    demo=400_000, flood_site_6=1_250_000, flood_site_8=1_450_000,
    podium_gsf=250, parking_gsf=115, office_gsf=290,
    office_ti_rsf=135, ground_ti_rsf=106.25,
    l2_port=7_500, ev_roughin_port=1_250,
    solar_w=2.40, bess_kwh_small=600, bess_kwh_large=550,
    roofgarden_sf=90,
)
# S2 alternative unit costs (for the dual-basis cost comparison)
U2 = dict(parking_gsf=105, office_gsf=310, office_ti_rsf=75, demo=850_000)

SOFT = dict(ae=0.08, permits=0.05, owner=0.04, contingency=0.08, leasing_rsf=25)
FIN = dict(ltc=0.60, rate=0.08, avg=0.55, fee=0.015)

REV = dict(office_rent=50, office_occ=0.93, ground_rent=45, ground_occ=0.95,
           data_rent=55, park_mo=275, monetized=0.55, ev_mo=250,
           opex_re=0.25, opex_park=0.35)
CITY_PARK_DIVISOR = 250     # ZON-04: 1 space / 250 SF GFA
TARGET_YOC = 0.07           # S3 institutional planning hurdle
EXIT_CAPS = dict(mob=0.0625, specialty=0.0725)   # OQ-18 unresolved -> report both


def k_factor(months):
    """TDC = k * pre-finance cost. Verified to reproduce S3 D56/E56 exactly."""
    return 1 + FIN["ltc"] * (FIN["rate"] * (months / 12) * FIN["avg"] + FIN["fee"])


def budget(sc):
    """Line-item planning budget. Returns dict of every cost line."""
    g = sc["geom"]
    b = {}
    b["Land purchase"] = sc["land"]
    b["Closing / acquisition costs"] = sc["land"] * CLOSING

    b["Demolition and disposal"] = U["demo"]
    b["Environmental remediation allowance"] = sc.get("remediation", 250_000)
    b["Flood / stormwater / resilient sitework"] = sc["flood_site"]
    b["Site paving, curb, lighting, landscape, streetscape"] = sc.get("sitework_surface", 0)
    b["Ground podium + active frontage shell"] = g["ground_gsf"] * U["podium_gsf"] if g["structured"] else g["ground_gsf"] * sc.get("single_story_gsf_cost", 340)
    b["Structured parking concrete + flat decks"] = g["park_gsf"] * U["parking_gsf"]
    b["Office / medical-ready shell and core"] = g["office_gsf"] * U["office_gsf"]
    b["Installed networked Level 2 chargers"] = sc["l2_installed"] * U["l2_port"]
    b["EV-ready expansion conduit / pathways"] = max(0, sc["ev_ready"] - sc["l2_installed"]) * U["ev_roughin_port"]
    b["FPL service / transformer / switchgear allowance"] = sc["fpl"]
    b["Mobility circulation, gates, security, staging"] = sc["av_fitout"]
    b["Rooftop / canopy solar"] = sc["solar_kw"] * 1000 * (sc.get("solar_cost_w") or U["solar_w"])
    b["Battery energy storage system"] = sc["bess_kwh"] * (U["bess_kwh_large"] if sc["bess_kwh"] >= 1000 else U["bess_kwh_small"])
    b["Edge data / teleoperations infrastructure"] = sc["data_infra"]
    b["Staff rooftop garden / shade amenity"] = sc["garden_sf"] * U["roofgarden_sf"]
    b["_CORE HARD SUBTOTAL"] = sum(v for kk, v in b.items() if not kk.startswith("_") and kk not in ("Land purchase", "Closing / acquisition costs"))

    b["Office / clinical TI allowance"] = g["office_rsf"] * U["office_ti_rsf"]
    b["Ground med-tail / cafe TI allowance"] = g["ground_rsf"] * U["ground_ti_rsf"]
    b["_TI SUBTOTAL"] = b["Office / clinical TI allowance"] + b["Ground med-tail / cafe TI allowance"]
    b["_FULL HARD"] = b["_CORE HARD SUBTOTAL"] + b["_TI SUBTOTAL"]

    for basis, hard in (("core", b["_CORE HARD SUBTOTAL"]), ("full", b["_FULL HARD"])):
        b[f"A&E ({basis} basis)"] = hard * SOFT["ae"]
        b[f"Permits / impact / utility fees ({basis} basis)"] = hard * SOFT["permits"]
        b[f"Developer / legal / insurance / owner ({basis} basis)"] = hard * SOFT["owner"]
        b[f"Construction contingency ({basis} basis)"] = hard * SOFT["contingency"]
    b["Leasing / commissioning / pre-opening"] = g["office_rsf"] * SOFT["leasing_rsf"]
    b["_CORE SOFT SUBTOTAL"] = sum(b[f"{x} (core basis)"] for x in
        ["A&E", "Permits / impact / utility fees", "Developer / legal / insurance / owner", "Construction contingency"])
    b["_FULL SOFT SUBTOTAL"] = sum(b[f"{x} (full basis)"] for x in
        ["A&E", "Permits / impact / utility fees", "Developer / legal / insurance / owner", "Construction contingency"]) + b["Leasing / commissioning / pre-opening"]

    kf = k_factor(sc["months"])
    b["_CORE PREFINANCE"] = b["Land purchase"] + b["Closing / acquisition costs"] + b["_CORE HARD SUBTOTAL"] + b["_CORE SOFT SUBTOTAL"]
    b["_FULL PREFINANCE"] = b["Land purchase"] + b["Closing / acquisition costs"] + b["_FULL HARD"] + b["_FULL SOFT SUBTOTAL"]
    b["_CORE TDC"] = b["_CORE PREFINANCE"] * kf
    b["_FULL TDC"] = b["_FULL PREFINANCE"] * kf
    b["Construction interest + financing fees (core)"] = b["_CORE TDC"] - b["_CORE PREFINANCE"]
    b["Construction interest + financing fees (tenant-ready)"] = b["_FULL TDC"] - b["_FULL PREFINANCE"]
    return b


def operating(sc, include_av=False):
    g = sc["geom"]
    r = {}
    r["Office / clinical EGI"] = g["office_rsf"] * REV["office_rent"] * REV["office_occ"]
    r["Ground med-tail / cafe EGI"] = g["ground_rsf"] * REV["ground_rent"] * REV["ground_occ"]
    r["Edge / data room EGI"] = sc["data_sf"] * REV["data_rent"] * REV["office_occ"]
    r["Monetized parking EGI"] = sc["monetizable_stalls"] * REV["park_mo"] * 12
    r["EV charging net revenue"] = sc["l2_installed"] * REV["ev_mo"] * 12
    r["AV staging / secure bay revenue"] = (sc["av_bays"] * 900 * 12) if include_av else 0.0
    r["_EGI"] = sum(v for kk, v in r.items() if not kk.startswith("_"))
    re_egi = r["Office / clinical EGI"] + r["Ground med-tail / cafe EGI"] + r["Edge / data room EGI"]
    mob_egi = r["Monetized parking EGI"] + r["EV charging net revenue"] + r["AV staging / secure bay revenue"]
    r["Real-estate operating expense"] = -re_egi * REV["opex_re"]
    r["Parking / EV / mobility operating expense"] = -mob_egi * REV["opex_park"]
    r["_NOI"] = r["_EGI"] + r["Real-estate operating expense"] + r["Parking / EV / mobility operating expense"]
    return r


def residual_land(noi, hard_soft, months, yoc):
    """Max supportable land at a stated yield on cost. Negative => no price works."""
    kf = k_factor(months)
    return (noi / (yoc * kf) - hard_soft) / (1 + CLOSING)


def stall_waterfall(sc):
    g = sc["geom"]
    gross = sc["structured_stalls"] + sc["grade_stalls"]
    code_office = math.ceil(g["office_gsf"] / CITY_PARK_DIVISOR)
    code_ground = math.ceil(g["ground_rsf"] / CITY_PARK_DIVISOR)
    code_total = code_office + code_ground
    ada = 0 if gross == 0 else (max(1, math.ceil(gross * 0.02)) if gross > 100 else max(1, math.ceil(gross / 25)))
    secure_loss = round(sc["gated_levels"] * STALLS_PER_LEVEL * 0.06)
    turnaround = sc["turnaround_equiv"]
    maneuver = sc["maneuver_reserve"]
    # A structured-deck shortfall cannot be offset against ground-level staging bays:
    # the bays sit on L1, are not counted toward the code parking requirement, and exist
    # whether or not the decks carry a surplus. Floor the structured figure at zero so a
    # deck deficit is reported as "no structured surplus" rather than cancelling the bays.
    net_fleet_raw = gross - code_total - secure_loss - turnaround - maneuver
    net_fleet = max(0, net_fleet_raw)
    return dict(gross_stalls=gross, code_required_office=code_office, code_required_ground=code_ground,
                code_required_total=code_total, ada_within_code=ada,
                secure_zone_boundary_loss=secure_loss, cleaning_turnaround_equivalent=turnaround,
                av_maneuvering_reserve=maneuver,
                net_operational_fleet_stalls=net_fleet,
                net_operational_fleet_stalls_before_floor=net_fleet_raw,
                plus_ground_av_bays=sc["av_bays"],
                total_operational_fleet_or_staging=net_fleet + sc["av_bays"],
                surplus_or_deficit_vs_code=gross - code_total)


def electrical(sc):
    """Load build-up. Method preserved from the corrected electrical workbook;
    depot load included only where the scenario places a depot on site."""
    g = sc["geom"]
    office_conn = g["office_gsf"] * 18 / 1000       # W/SF clinical-capable
    medtail_conn = g["ground_rsf"] * 22 / 1000
    park_conn = (g["park_gsf"] + g["ground_gsf"] - g["ground_rsf"]) * 1.5 / 1000
    misc = sc.get("misc_kw", 100)
    imaging = sc.get("imaging_kw", 0)
    base = office_conn * 0.7 + medtail_conn * 0.8 + park_conn * 0.8 + misc + imaging
    depot_peak = 0.0
    if sc.get("depot_vehicles", 0):
        daily_kwh = sc["depot_vehicles"] * 275 * 0.35
        depot_peak = daily_kwh / sc.get("charge_window_hr", 8)
    l2 = sc["l2_installed"] * 9.6 * 0.5
    dcfc = sc.get("dcfc_units", 0) * 150 * 0.6
    charging = depot_peak + l2 + dcfc
    unmanaged = base * 0.9 + charging * 1.25
    managed = unmanaged - sc.get("bess_shave_kw", 0)
    amps = managed * 1000 / (480 * math.sqrt(3) * 0.9)
    return dict(base_connected_kw=round(office_conn + medtail_conn + park_conn + misc + imaging, 1),
                base_demand_kw=round(base, 1),
                depot_managed_peak_kw=round(depot_peak, 1),
                public_l2_demand_kw=round(l2, 1), public_dcfc_demand_kw=round(dcfc, 1),
                total_charging_demand_kw=round(charging, 1),
                service_basis_unmanaged_kw=round(unmanaged, 1),
                service_basis_managed_kw=round(managed, 1),
                amps_at_480v_managed=round(amps),
                service_class="MEDIUM VOLTAGE (13.2 kV) indicated" if managed > 2500 else "480 V secondary feasible",
                recommended_service_mva=round(managed * 1.25 / 1000, 2))


def geom(ground_gsf, ground_rsf, park_levels, office_levels, structured=True, office_plate=PLATE_OFFICE):
    park_gsf = park_levels * PLATE_PARK
    office_gsf = office_levels * office_plate
    return dict(ground_gsf=ground_gsf, ground_rsf=ground_rsf, park_gsf=park_gsf,
                office_gsf=office_gsf, office_rsf=office_gsf * OFFICE_EFF,
                total_gsf=ground_gsf + park_gsf + office_gsf,
                park_levels=park_levels, office_levels=office_levels, structured=structured)


# ---------------------------------------------------------------- scenarios
S = {}

S["SA-A"] = dict(
    name="Maximum-Fleet Healthcare Mobility Hub (8 storeys, on-site depot)",
    stories=8, geom=geom(28_000, 10_000, 5, 2), land=LAND_BASE, months=34,
    flood_site=U["flood_site_8"], l2_installed=40, ev_ready=120, dcfc_units=8,
    fpl=3_500_000, av_fitout=2_400_000, solar_kw=200, solar_cost_w=4.00, bess_kwh=2000,
    data_infra=1_750_000, data_sf=2000, garden_sf=8000,
    structured_stalls=5 * STALLS_PER_LEVEL, grade_stalls=0, gated_levels=2,
    turnaround_equiv=12, maneuver_reserve=10, av_bays=18,
    depot_vehicles=120, charge_window_hr=8, bess_shave_kw=700,
    monetizable_stalls=0,  # set below
)

S["SA-B"] = dict(
    name="Balanced Clinical and Staging Hub — two-site strategy (6 storeys)",
    stories=6, geom=geom(28_000, 10_000, 3, 2), land=LAND_BASE, months=30,
    flood_site=U["flood_site_6"], l2_installed=40, ev_ready=80, dcfc_units=0,
    fpl=1_500_000, av_fitout=1_250_000, solar_kw=150, bess_kwh=500,
    data_infra=1_250_000, data_sf=1500, garden_sf=6000,
    structured_stalls=3 * STALLS_PER_LEVEL, grade_stalls=0, gated_levels=1,
    turnaround_equiv=6, maneuver_reserve=6, av_bays=18,
    depot_vehicles=0, bess_shave_kw=0, monetizable_stalls=0,
)

S["SA-C"] = dict(
    name="Minimum Viable Development (4 storeys, expansion-ready)",
    stories=4, geom=geom(28_000, 12_000, 2, 1), land=LAND_BASE, months=22,
    flood_site=1_150_000, l2_installed=40, ev_ready=60, dcfc_units=0,
    fpl=900_000, av_fitout=750_000, solar_kw=100, bess_kwh=0,
    data_infra=0, data_sf=0, garden_sf=0,
    structured_stalls=2 * STALLS_PER_LEVEL, grade_stalls=15, gated_levels=0,
    turnaround_equiv=4, maneuver_reserve=4, av_bays=10,
    depot_vehicles=0, bess_shave_kw=0, monetizable_stalls=0,
)

S["SA-C0"] = dict(
    name="Implementation Staging Ground (single-storey med-tail + surface mobility yard)",
    stories=1, geom=geom(14_000, 14_000, 0, 0, structured=False), land=LAND_BASE, months=16,
    flood_site=900_000, sitework_surface=850_000, single_story_gsf_cost=340,
    l2_installed=40, ev_ready=70, dcfc_units=0,
    fpl=600_000, av_fitout=600_000, solar_kw=100, solar_cost_w=4.00, bess_kwh=0,
    data_infra=0, data_sf=0, garden_sf=0,
    structured_stalls=0, grade_stalls=70, gated_levels=0,
    turnaround_equiv=3, maneuver_reserve=3, av_bays=8,
    depot_vehicles=0, bess_shave_kw=0, monetizable_stalls=0,
)

# monetizable stalls: S3 construct (55% of structured) BUT capped by genuine surplus over code.
for key, sc in S.items():
    w = stall_waterfall(sc)
    s3_construct = round((sc["structured_stalls"]) * REV["monetized"])
    code_constrained = max(0, w["surplus_or_deficit_vs_code"])
    sc["_waterfall"] = w
    sc["_monetized_s3_construct"] = s3_construct
    sc["_monetized_code_constrained"] = code_constrained
    sc["monetizable_stalls"] = code_constrained          # governance-clean: never sell code-required stalls

results = {}
for key, sc in S.items():
    b = budget(sc)
    op_clean = operating(sc, include_av=False)
    op_av = operating(sc, include_av=True)
    # Parking construct P2: S3's "55% of structured stalls monetized" — reported alongside
    # the code-constrained construct P1 so the range is visible rather than assumed away.
    sc_p2 = dict(sc); sc_p2["monetizable_stalls"] = sc["_monetized_s3_construct"]
    op_p2 = operating(sc_p2, include_av=False)
    g = sc["geom"]
    core_hs = b["_CORE HARD SUBTOTAL"] + b["_CORE SOFT SUBTOTAL"]
    full_hs = b["_FULL HARD"] + b["_FULL SOFT SUBTOTAL"]
    noi = op_clean["_NOI"]
    r = dict(
        scenario=key, name=sc["name"], stories=sc["stories"],
        geometry={kk: (round(vv) if isinstance(vv, float) else vv) for kk, vv in g.items()},
        stall_waterfall=sc["_waterfall"],
        monetized_stalls_s3_construct=sc["_monetized_s3_construct"],
        monetized_stalls_code_constrained=sc["_monetized_code_constrained"],
        budget={kk: round(vv) for kk, vv in b.items()},
        operating_governance_clean={kk: round(vv) for kk, vv in op_clean.items()},
        operating_with_av_upside={kk: round(vv) for kk, vv in op_av.items()},
        operating_p2_s3_parking_construct={kk: round(vv) for kk, vv in op_p2.items()},
        metrics=dict(
            core_all_in=round(b["_CORE TDC"]),
            tenant_ready_all_in=round(b["_FULL TDC"]),
            core_hard=round(b["_CORE HARD SUBTOTAL"]),
            core_soft=round(b["_CORE SOFT SUBTOTAL"]),
            cost_ex_land_core=round(b["_CORE TDC"] - b["Land purchase"]),
            cost_per_gsf_core=round(b["_CORE TDC"] / g["total_gsf"], 2) if g["total_gsf"] else None,
            cost_per_gross_stall=round(b["Structured parking concrete + flat decks"] / sc["structured_stalls"]) if sc["structured_stalls"] else None,
            cost_per_operational_fleet_stall=round(b["_CORE TDC"] / sc["_waterfall"]["total_operational_fleet_or_staging"]) if sc["_waterfall"]["total_operational_fleet_or_staging"] > 0 else None,
            cost_per_clinical_rsf=round(b["_FULL TDC"] / (g["office_rsf"] + g["ground_rsf"]), 2) if (g["office_rsf"] + g["ground_rsf"]) else None,
            noi_governance_clean=round(noi),
            noi_with_av_upside=round(op_av["_NOI"]),
            yoc_core=round(noi / b["_CORE TDC"], 5),
            yoc_tenant_ready=round(noi / b["_FULL TDC"], 5),
            value_at_cap_625=round(noi / EXIT_CAPS["mob"]),
            value_at_cap_725=round(noi / EXIT_CAPS["specialty"]),
            value_gap_core_at_725=round(noi / EXIT_CAPS["specialty"] - b["_CORE TDC"]),
            value_gap_tenant_ready_at_725=round(noi / EXIT_CAPS["specialty"] - b["_FULL TDC"]),
            annual_partner_gap_core=round(max(0, b["_CORE TDC"] * TARGET_YOC - noi)),
            annual_partner_gap_tenant_ready=round(max(0, b["_FULL TDC"] * TARGET_YOC - noi)),
            noi_p2_s3_parking_construct=round(op_p2["_NOI"]),
            yoc_core_p2_s3_parking_construct=round(op_p2["_NOI"] / b["_CORE TDC"], 5),
            noi_p2_plus_av_upside=round(op_p2["_NOI"] + (sc["av_bays"] * 900 * 12) * (1 - REV["opex_park"])),
            yoc_core_p2_plus_av=round((op_p2["_NOI"] + (sc["av_bays"] * 900 * 12) * (1 - REV["opex_park"])) / b["_CORE TDC"], 5),
        ),
        residual_land_value={
            f"core_at_{int(y*1000)}bps": round(residual_land(noi, core_hs, sc["months"], y))
            for y in (0.07, 0.065, 0.06, 0.055, 0.05, 0.045, 0.04)
        },
        residual_land_value_tenant_ready={
            f"tenant_ready_at_{int(y*1000)}bps": round(residual_land(noi, full_hs, sc["months"], y))
            for y in (0.07, 0.06, 0.05, 0.04)
        },
        electrical=electrical(sc),
    )
    # sensitivity: hard/soft/rate factors (S3 Sensitivity construct)
    sens = {}
    for label, hf, sf, rate in (("low", 0.90, 0.90, 0.070), ("base", 1.00, 1.00, 0.080), ("high", 1.12, 1.15, 0.095)):
        pre = b["Land purchase"] + b["Closing / acquisition costs"] + b["_CORE HARD SUBTOTAL"] * hf + b["_CORE SOFT SUBTOTAL"] * sf
        kf = 1 + FIN["ltc"] * (rate * (sc["months"] / 12) * FIN["avg"] + FIN["fee"])
        sens[label] = dict(core_all_in=round(pre * kf), yoc_core=round(noi / (pre * kf), 5))
    r["sensitivity_cost"] = sens
    # break-even office rent to reach 7% YoC on core
    other_noi = noi - (g["office_rsf"] * REV["office_rent"] * REV["office_occ"]) * (1 - REV["opex_re"])
    need = b["_CORE TDC"] * TARGET_YOC
    r["breakeven"] = dict(
        office_rent_for_7pct_yoc_core=round((need - other_noi) / (g["office_rsf"] * REV["office_occ"] * (1 - REV["opex_re"])), 2) if g["office_rsf"] else None,
        land_price_for_7pct_yoc_core=round(residual_land(noi, core_hs, sc["months"], 0.07)),
        noi_required_for_7pct_yoc_core=round(need),
        noi_shortfall=round(need - noi),
    )
    results[key] = r

# ---- Site Enablement Package: minimum spend to make the site function as a staging ground
enable = {
    "Demolition and disposal": U["demo"],
    "Environmental remediation allowance": 250_000,
    "Flood / stormwater / drainage / exfiltration": 900_000,
    "Paving, curb, striping, fencing, lighting": 650_000,
    "Electrical service, panel, distribution": 600_000,
    "40 networked Level 2 chargers + conduit for 30 more": 40 * U["l2_port"] + 30 * U["ev_roughin_port"],
    "Passenger-transfer canopy, gates, LPR, wayfinding": 600_000,
    "Streetscape / landscape / screening (City frontage condition)": 350_000,
}
enable["_HARD SUBTOTAL"] = sum(v for kk, v in enable.items() if not kk.startswith("_"))
enable["_SOFT @22%"] = enable["_HARD SUBTOTAL"] * 0.22
enable["_TOTAL EX-LAND"] = enable["_HARD SUBTOTAL"] + enable["_SOFT @22%"]

# ---- Stranded-cost exposure for SA-B if the off-site depot never lands
b_b = results["SA-B"]["budget"]
stranded = {
    "Mobility circulation, gates, security, staging": b_b["Mobility circulation, gates, security, staging"],
    "Edge data / teleoperations infrastructure": b_b["Edge data / teleoperations infrastructure"],
    "EV-ready expansion conduit beyond installed": b_b["EV-ready expansion conduit / pathways"],
    "Incremental FPL service above a clinical-only requirement": 1_500_000 - 900_000,
}
stranded["_TOTAL AT RISK"] = sum(stranded.values())

# ---- Two-site comparison: on-site depot vs off-site depot
depot_onsite_gsf = 55_000
two_site = {
    "On-site depot structured area (GSF)": depot_onsite_gsf,
    "On-site depot shell cost at parking rate ($115/GSF)": depot_onsite_gsf * U["parking_gsf"],
    "On-site depot shell cost at ventilated podium rate ($250/GSF)": depot_onsite_gsf * U["podium_gsf"],
    "Implied South Andrews land allocated to depot (GSF/plate x $209/SF land)": round(depot_onsite_gsf / 28_000 * 28_000 / 38_207 * LAND_BASE),
    "SA-A managed service basis (kW)": results["SA-A"]["electrical"]["service_basis_managed_kw"],
    "SA-B managed service basis (kW)": results["SA-B"]["electrical"]["service_basis_managed_kw"],
    "Service cost, medium-voltage case (electrical workbook JUDGMENT)": 3_500_000,
    "Service cost, 480V clinical case (S3 allowance)": 1_500_000,
    "BESS required to hold MV case down": 2000 * U["bess_kwh_large"],
    "Off-site depot land, 2.5 acres at $40/SF (illustrative benchmark, unverified)": round(2.5 * 43560 * 40),
}

payload = dict(
    disclaimer=("PLANNING-LEVEL MODEL OUTPUT. Derived from S2/S3 workbook inputs and stated assumptions. "
                "Not an appraisal, contractor estimate, engineered load letter, tax opinion, lease forecast, "
                "or financing commitment. Geometric basis: Basis B (28k/24k plates, 360 GSF/stall) — a stated "
                "modeling basis, not an adopted program (OQ-14). AV revenue excluded from every base case per CAN-012."),
    stated_inputs=dict(site_sf=SITE_SF, land_working_input=LAND_BASE, closing=CLOSING,
                       plates=dict(ground=PLATE_GROUND, parking=PLATE_PARK, office=PLATE_OFFICE),
                       gsf_per_stall=GSF_PER_STALL, stalls_per_level=STALLS_PER_LEVEL,
                       office_efficiency=OFFICE_EFF, unit_costs=U, s2_alternative_unit_costs=U2,
                       soft=SOFT, financing=FIN, revenue=REV,
                       city_parking_divisor=CITY_PARK_DIVISOR, target_yoc=TARGET_YOC, exit_caps=EXIT_CAPS),
    scenarios=results,
    site_enablement_package={kk: round(vv) for kk, vv in enable.items()},
    stranded_cost_exposure_SA_B={kk: round(vv) for kk, vv in stranded.items()},
    two_site_comparison=two_site,
)

with open(os.path.join(OUT, "integrated-development-model.json"), "w") as f:
    json.dump(payload, f, indent=2)
    f.write("\n")


# ---------------------------------------------------------------- derived workbook
def write_workbook():
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment
    except ImportError:
        print("openpyxl unavailable — JSON written, workbook skipped")
        return
    wb = Workbook()
    bold = Font(bold=True)
    wrap = Alignment(wrap_text=True, vertical="top")

    ws = wb.active
    ws.title = "Basis and Limitations"
    for row in [
        ["SOUTH ANDREWS HEALTHCARE AND MOBILITY HUB — INTEGRATED DEVELOPMENT MODEL (DERIVED)"],
        ["901–917 South Andrews Avenue, Fort Lauderdale, Florida"],
        [],
        ["STATUS", "Noncanonical working model. NOT the gated model vNext (Document Update Order Step 5)."],
        ["CLASSIFICATION", "model output — planning-level. Not an appraisal, contractor estimate, engineered load letter, tax opinion, lease forecast, or financing commitment."],
        ["PROPERTY STATUS", "Prospective acquisition. The sponsor does not currently own the property."],
        [],
        ["GEOMETRIC BASIS", "Basis B (S3 Assumptions!D9:D12, from S8 pp.4-5). 28,000 GSF ground/parking plate; 24,000 GSF office plate; 360 GSF/stall; 85% office efficiency. STATED BASIS ONLY — not adopted (OQ-14). Basis A (S2, 35,000 SF plate) changes every figure. Never mix bases."],
        ["UNIT COSTS", "S3 Assumptions!D29:D43 (2026 WGI-benchmarked). S2 alternatives shown separately on 'Stated Inputs'."],
        ["REVENUE", "S3 Assumptions!D55:D65. Office rent $50/RSF is an illustrative broker-underwriting input, not a signed lease (OQ-15)."],
        ["AV REVENUE", "Excluded from every base case per AGENTS.md and CAN-012. Shown only as separately labelled upside."],
        ["TAX CREDITS", "Solar/storage underwritten at 0% per S3 Energy & Mobility!B21 pending tax counsel (OQ-17 / MB-13)."],
        ["LAND BASIS", "$8,000,000 is the D-P1 working opening-offer input — an acquisition-strategy input only, not proof of value, not a price ceiling, walk-away price, or transaction authority."],
        ["ELECTRICAL", "Method preserved from the corrected electrical load workbook, which is NOT held in sources/ (MA-11). Its 18 W/SF clinical density conflicts with S3's 5 W/GSF; both reported."],
        [],
        ["VALIDATION", "SA-B on parking construct P2 with AV re-added reproduces S3 Financing & Returns!D17 ($2,277,951) within $107."],
        ["PARKING CONSTRUCTS", "P1 = only stalls surplus to the City 1-per-250-GFA standard are monetized (governance-clean). P2 = S3's 55%-of-structured-stalls construct. Both reported; neither adopted pending MB-01 and MB-09."],
    ]:
        ws.append(row)
    ws["A1"].font = Font(bold=True, size=13)
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 130
    for r in ws.iter_rows(min_col=2, max_col=2):
        for c in r:
            c.alignment = wrap

    ws = wb.create_sheet("Scenario Comparison")
    hdr = ["Metric"] + [f"{k} — {results[k]['stories']} storeys" for k in results]
    ws.append(hdr)
    def row(label, fn):
        ws.append([label] + [fn(results[k]) for k in results])
    row("Total constructed GSF", lambda r: r["geometry"]["total_gsf"])
    row("Office/clinical RSF", lambda r: r["geometry"]["office_rsf"])
    row("Ground leasable RSF", lambda r: r["geometry"]["ground_rsf"])
    row("Gross stalls", lambda r: r["stall_waterfall"]["gross_stalls"])
    row("City code-required stalls (1/250 GFA)", lambda r: r["stall_waterfall"]["code_required_total"])
    row("Stall surplus / (deficit) vs code", lambda r: r["stall_waterfall"]["surplus_or_deficit_vs_code"])
    row("Net operational fleet/staging positions", lambda r: r["stall_waterfall"]["total_operational_fleet_or_staging"])
    row("Core hard cost", lambda r: r["metrics"]["core_hard"])
    row("Core soft cost", lambda r: r["metrics"]["core_soft"])
    row("Core all-in cost (incl. land)", lambda r: r["metrics"]["core_all_in"])
    row("Tenant-ready all-in cost", lambda r: r["metrics"]["tenant_ready_all_in"])
    row("Cost excluding land (core)", lambda r: r["metrics"]["cost_ex_land_core"])
    row("Core all-in $/GSF", lambda r: r["metrics"]["cost_per_gsf_core"])
    row("Parking structure $/gross stall", lambda r: r["metrics"]["cost_per_gross_stall"])
    row("Core all-in $/operational fleet position", lambda r: r["metrics"]["cost_per_operational_fleet_stall"])
    row("Tenant-ready $/clinical+ground RSF", lambda r: r["metrics"]["cost_per_clinical_rsf"])
    row("NOI — P1 code-constrained parking, no AV", lambda r: r["metrics"]["noi_governance_clean"])
    row("NOI — P2 S3 parking construct, no AV", lambda r: r["metrics"]["noi_p2_s3_parking_construct"])
    row("Yield on cost — core, P1", lambda r: r["metrics"]["yoc_core"])
    row("Yield on cost — core, P2", lambda r: r["metrics"]["yoc_core_p2_s3_parking_construct"])
    row("Yield on cost — core, P2 + AV upside", lambda r: r["metrics"]["yoc_core_p2_plus_av"])
    row("Annual partner/grant gap to 7% YoC (core)", lambda r: r["metrics"]["annual_partner_gap_core"])
    row("Value gap vs core cost at 7.25% cap", lambda r: r["metrics"]["value_gap_core_at_725"])
    row("Break-even office rent for 7% YoC ($/RSF)", lambda r: r["breakeven"]["office_rent_for_7pct_yoc_core"])
    row("Max rational land price @ 7.0% YoC", lambda r: r["residual_land_value"]["core_at_70bps"])
    row("Max rational land price @ 6.0% YoC", lambda r: r["residual_land_value"]["core_at_60bps"])
    row("Max rational land price @ 5.0% YoC", lambda r: r["residual_land_value"]["core_at_50bps"])
    row("Max rational land price @ 4.0% YoC", lambda r: r["residual_land_value"]["core_at_40bps"])
    row("Managed electrical service basis (kW)", lambda r: r["electrical"]["service_basis_managed_kw"])
    row("Indicated service class", lambda r: r["electrical"]["service_class"])
    for c in ws[1]:
        c.font = bold
    ws.column_dimensions["A"].width = 46
    for col in "BCDE":
        ws.column_dimensions[col].width = 24

    for key, r in results.items():
        ws = wb.create_sheet(f"{key} Budget"[:31])
        ws.append([f"{key} — {r['name']}"]); ws["A1"].font = bold
        ws.append([])
        ws.append(["Line item", "Amount", "Basis / provenance"])
        for c in ws[3]:
            c.font = bold
        prov = {
            "Land purchase": "D-P1 working opening-offer input; not value, ceiling or authority",
            "Closing / acquisition costs": "S3 Assumptions!D8 — 2% of land",
            "Demolition and disposal": "S3 Assumptions!D29 (S2 B19 carries $850,000 incl. sitework)",
            "Environmental remediation allowance": "Analyst allowance — Phase I/II exist per S6 p.10 but are not in the repository (MA-13)",
            "Flood / stormwater / resilient sitework": "S3 Assumptions!D30/E30 — FEMA Zone AE",
            "Ground podium + active frontage shell": "S3 Assumptions!D31 — $250/GSF",
            "Structured parking concrete + flat decks": "S3 Assumptions!D32 — $115/GSF (WGI 2026 median $98.75 + S FL/HVHZ premium)",
            "Office / medical-ready shell and core": "S3 Assumptions!D33 — $290/GSF (S2 B25 carries $310)",
            "Installed networked Level 2 chargers": "S3 Assumptions!D36 — $7,500/port",
            "EV-ready expansion conduit / pathways": "S3 Assumptions!D37 — $1,250/port",
            "FPL service / transformer / switchgear allowance": "S3 Assumptions!D38 high-uncertainty placeholder; MV cases use the corrected electrical workbook's JUDGMENT figure. Subject to utility confirmation (MB-03)",
            "Mobility circulation, gates, security, staging": "S3 Assumptions!D39",
            "Rooftop / canopy solar": "S3 Assumptions!D40 $2.40/W rooftop; $4.00/W where canopy-mounted per corrected electrical workbook",
            "Battery energy storage system": "S3 Assumptions!D41",
            "Edge data / teleoperations infrastructure": "S3 Assumptions!D42",
            "Staff rooftop garden / shade amenity": "S3 Assumptions!D43 — $90/SF",
            "Office / clinical TI allowance": "S3 Assumptions!D34 — $135/RSF (S2 B26 carries $75; OQ scope difference)",
            "Ground med-tail / cafe TI allowance": "S3 Assumptions!D35 — $106.25/RSF",
            "Leasing / commissioning / pre-opening": "S3 Assumptions!D49 — $25/office RSF",
        }
        for kk, vv in r["budget"].items():
            if kk.startswith("_"):
                ws.append([kk.strip("_").title(), vv, "subtotal"])
                ws.cell(ws.max_row, 1).font = bold
                ws.cell(ws.max_row, 2).font = bold
            else:
                ws.append([kk, vv, prov.get(kk, "S3 Assumptions / analyst allowance")])
        ws.append([])
        ws.append(["OPERATING — P1 code-constrained parking, AV excluded"]); ws.cell(ws.max_row, 1).font = bold
        for kk, vv in r["operating_governance_clean"].items():
            ws.append([kk.strip("_"), vv])
        ws.append([])
        ws.append(["OPERATING — P2 S3 parking construct, AV excluded"]); ws.cell(ws.max_row, 1).font = bold
        for kk, vv in r["operating_p2_s3_parking_construct"].items():
            ws.append([kk.strip("_"), vv])
        ws.append([])
        ws.append(["STALL WATERFALL — gross stalls are not operational fleet capacity"]); ws.cell(ws.max_row, 1).font = bold
        for kk, vv in r["stall_waterfall"].items():
            ws.append([kk.replace("_", " "), vv])
        ws.append([])
        ws.append(["ELECTRICAL LOAD BUILD-UP"]); ws.cell(ws.max_row, 1).font = bold
        for kk, vv in r["electrical"].items():
            ws.append([kk.replace("_", " "), vv])
        ws.append([])
        ws.append(["MAX RATIONAL LAND PRICE BY TARGET YIELD ON COST"]); ws.cell(ws.max_row, 1).font = bold
        for kk, vv in r["residual_land_value"].items():
            ws.append([kk.replace("_", " "), vv])
        ws.append([])
        ws.append(["COST SENSITIVITY (hard / soft / rate factors per S3 Sensitivity!A6:D8)"]); ws.cell(ws.max_row, 1).font = bold
        for kk, vv in r["sensitivity_cost"].items():
            ws.append([kk, vv["core_all_in"], f"YoC {vv['yoc_core']*100:.2f}% (P1)"])
        ws.column_dimensions["A"].width = 52
        ws.column_dimensions["B"].width = 18
        ws.column_dimensions["C"].width = 96
        for rr in ws.iter_rows(min_col=3, max_col=3):
            for c in rr:
                c.alignment = wrap

    ws = wb.create_sheet("Enablement and Two-Site")
    ws.append(["SITE ENABLEMENT PACKAGE — minimum spend to make the site function as a staging ground (excludes land)"])
    ws["A1"].font = bold
    for kk, vv in payload["site_enablement_package"].items():
        ws.append([kk.strip("_"), vv])
    ws.append([])
    ws.append(["STRANDED-COST EXPOSURE — SA-B if the off-site depot is never secured"]); ws.cell(ws.max_row, 1).font = bold
    for kk, vv in payload["stranded_cost_exposure_SA_B"].items():
        ws.append([kk.strip("_"), vv])
    ws.append([])
    ws.append(["TWO-SITE COMPARISON"]); ws.cell(ws.max_row, 1).font = bold
    for kk, vv in payload["two_site_comparison"].items():
        ws.append([kk, vv])
    ws.column_dimensions["A"].width = 76
    ws.column_dimensions["B"].width = 22

    ws = wb.create_sheet("Stated Inputs")
    ws.append(["Input group", "Key", "Value"]); [setattr(c, "font", bold) for c in ws[1]]
    for group, d in payload["stated_inputs"].items():
        if isinstance(d, dict):
            for kk, vv in d.items():
                ws.append([group, kk, str(vv)])
        else:
            ws.append([group, "", str(d)])
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 34
    ws.column_dimensions["C"].width = 60

    path = os.path.join(OUT, "South_Andrews_Integrated_Development_Model.xlsx")
    wb.save(path)
    print("wrote", os.path.basename(path))


write_workbook()

# ---------------------------------------------------------------- console report
def money(x): return f"${x:,.0f}"
print("=" * 104)
for key in ["SA-A", "SA-B", "SA-C", "SA-C0"]:
    r = results[key]; m = r["metrics"]; w = r["stall_waterfall"]; e = r["electrical"]
    print(f"\n### {key} — {r['name']}  [{r['stories']} storeys, {r['geometry']['total_gsf']:,} GSF]")
    print(f"  STALLS  gross {w['gross_stalls']} | code-required {w['code_required_total']} "
          f"| surplus/deficit {w['surplus_or_deficit_vs_code']:+d} | net operational fleet {w['net_operational_fleet_stalls']} "
          f"+ {w['plus_ground_av_bays']} bays = {w['total_operational_fleet_or_staging']}")
    print(f"  COST    core hard {money(m['core_hard'])} | core all-in {money(m['core_all_in'])} "
          f"| tenant-ready {money(m['tenant_ready_all_in'])} | {money(m['cost_per_gsf_core'])}/GSF")
    print(f"  NOI     governance-clean {money(m['noi_governance_clean'])} | YoC core {m['yoc_core']*100:.2f}% "
          f"| tenant-ready {m['yoc_tenant_ready']*100:.2f}%")
    print(f"  GAP     annual partner gap (core) {money(m['annual_partner_gap_core'])} "
          f"| value gap @7.25% cap {money(m['value_gap_core_at_725'])}")
    print(f"  LAND    max rational land @7.0% {money(r['residual_land_value']['core_at_70bps'])} "
          f"| @6.0% {money(r['residual_land_value']['core_at_60bps'])} "
          f"| @5.0% {money(r['residual_land_value']['core_at_50bps'])} "
          f"| @4.0% {money(r['residual_land_value']['core_at_40bps'])}")
    print(f"  ELEC    base demand {e['base_demand_kw']} kW | charging {e['total_charging_demand_kw']} kW "
          f"| managed service {e['service_basis_managed_kw']} kW -> {e['recommended_service_mva']} MVA | {e['service_class']}")
    print(f"  BE      office rent for 7% YoC: ${r['breakeven']['office_rent_for_7pct_yoc_core']}/RSF "
          f"(modeled at ${REV['office_rent']}) | NOI shortfall {money(r['breakeven']['noi_shortfall'])}")

print("\n" + "=" * 104)
print("SITE ENABLEMENT PACKAGE (ex-land):", money(enable["_TOTAL EX-LAND"]))
print("SA-B STRANDED IF NO DEPOT:", money(stranded["_TOTAL AT RISK"]))
print("\nTwo-site:", json.dumps({k: (money(v) if isinstance(v, (int, float)) and v > 10000 else v) for k, v in two_site.items()}, indent=2))
print("\nwrote integrated-development-model.json")
